"""
Mobile Parts Finder - scripts/import-xlsx.py
==============================================================================
Imports the multi-sheet brand/model workbook into data/raw/xlsx_models.json,
which is what scripts/build-dataset.js reads.

    python scripts/import-xlsx.py "C:/Users/stark/Downloads/All_Brands_Models.xlsx"
    python scripts/import-xlsx.py <file> --dry     # report only, write nothing

ONE SHEET IS ONE BRAND. The sheet name IS the brand, used verbatim, and a row
can only ever join the brand of the sheet it came from. Identity is
(brand, model title), so the same title under two brands is two devices and the
same title twice under one brand is one device.

COLUMNS ARE READ BY HEADER, NOT BY POSITION.
    The previous importer indexed the row tuple: r[3] was the model name, r[6]
    the height. That is fine until someone inserts a column in Excel, at which
    point every model silently acquires another model's height and nothing
    fails loudly. Headers are matched on a loosened form of their text, so
    "Screen Ratio (cm2)", "screen ratio (cm^2)" and the mojibake "cm?" that the
    current file actually contains all resolve to the same field.

NOTHING IS INVENTED.
    An empty cell imports as None. It is never a zero, never an empty string
    standing in for a number, never a guess from a similar model. The UI is
    what decides how to show an absent value; the database records that it is
    absent. A spec sheet that invents a battery capacity is worse than one that
    admits the field is unknown, because the invented one gets quoted to a
    customer.

    Nothing is fetched, either. No GSMArena, no scraping, no enrichment. The
    image and source links are stored exactly as the workbook gives them.

RE-RUNNING IS SAFE, AND EMPTY CELLS NEVER ERASE.
    The import merges into whatever is already stored. A model already in the
    file keeps every value it has unless this workbook carries a real value for
    that field. So importing a sparse sheet over a rich record adds to it and
    takes nothing away, and running the same workbook twice changes nothing at
    all the second time.
==============================================================================
"""
import sys, os, io, json, re, unicodedata

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "data", "raw", "xlsx_models.json")
REPORT = os.path.join(ROOT, "data", "raw", "import-report.json")

args = [a for a in sys.argv[1:] if not a.startswith("--")]
flags = set(a for a in sys.argv[1:] if a.startswith("--"))
SRC = args[0] if args else r"C:\Users\stark\Downloads\All_Brands_Models.xlsx"
DRY = "--dry" in flags


# ----------------------------------------------------------------- headers
def loose(s):
    """Header text reduced to something matchable.

    Drops case, punctuation, units and any character that is not ASCII
    alphanumeric. The workbook's own header reads 'Screen Ratio (cm\ufffd)' —
    the superscript two survived a bad encoding — and this is why matching on
    the exact string is not good enough.
    """
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = s.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "", s.lower())


# field -> the header spellings that mean it. First match wins.
COLUMNS = {
    "serial":  ["serialnumber", "serialno", "sno", "slno"],
    "gsm":     ["gsmarenasourcelink", "gsmarenalink", "sourcelink", "gsmarena"],
    "img":     ["imagelink", "imageurl", "image", "photo"],
    "name":    ["mobilemodelname", "modelname", "model", "mobilemodel", "name"],
    "release": ["releasedate", "released", "launchdate"],
    "size":    ["sizeinches", "displaysize", "screensize", "size"],
    "h":       ["heightmm", "height"],
    "w":       ["widthmm", "width"],
    "scr":     ["screenratiocm2", "screenratiocm", "screenareacm2", "screenratio", "screenarea"],
    "ratio":   ["bodytoscreenratio", "screentobodyratio", "bodyratio"],
    "mah":     ["batterycapacitymah", "batterycapacity", "batterymah", "battery"],
}
REQUIRED = "name"      # a row with no model title is not a model


def map_headers(header_row):
    """header cell index -> field name, by matching the loosened header text."""
    seen = {}
    for idx, cell in enumerate(header_row or ()):
        key = loose(cell)
        if not key:
            continue
        for field, spellings in COLUMNS.items():
            if field in seen:
                continue
            if key in spellings:
                seen[field] = idx
                break
    return seen


# ------------------------------------------------------------------ values
def text(v):
    """Trimmed string, or None. Whitespace-only is nothing, not an empty title."""
    if v is None:
        return None
    s = str(v).replace("\u00a0", " ").strip()
    s = re.sub(r"\s+", " ", s)          # collapse runs; the title is otherwise untouched
    return s or None


def number(v):
    """A real number, or None. Never 0 as a stand-in for 'not recorded'."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return v
    s = re.sub(r"[^0-9.\-]", "", str(v))
    if s in ("", "-", ".", "-."):
        return None
    try:
        f = float(s)
    except ValueError:
        return None
    return int(f) if f.is_integer() else f


def date_text(v):
    """Kept as text. build-dataset.js parses d/m/Y and ISO; both survive here."""
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%d/%m/%Y")
    return text(v)


CASTS = {"serial": number,
         "size": number, "h": number, "w": number, "scr": number,
         "ratio": number, "mah": number, "release": date_text}


def cast(field, raw):
    return CASTS.get(field, text)(raw)


# ------------------------------------------------------------------- merge
def ident(s):
    """Identity form of a brand or model title.

    Deliberately NOT the same as loose(). This mirrors slug() in
    scripts/build-dataset.js character for character, because a model that is
    one record here must be one record there.

    Two separators matter and neither may be discarded:

      "+"     Coolpad Cool 20 and Coolpad Cool 20+ are different phones, so it
              is spelled out rather than dropped.
      spaces  the Honor sheet carries BOTH "Honor Play9A" (2024, 6.56",
              5200 mAh) and "Honor Play 9A" (2020, 6.3", 5000 mAh). They are
              two devices with two GSMArena pages. Collapsing whitespace away
              merges them into one and silently deletes a phone, which is the
              exact merging this import must not do.

    So punctuation becomes a separator instead of vanishing.
    """
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = s.encode("ascii", "ignore").decode("ascii")
    s = s.lower().replace("+", " plus ")
    return re.sub(r"[^a-z0-9]+", "-", s).strip("-")


def key_of(brand, name):
    """Identity. Brand is part of it, so a title cannot drift across brands."""
    return ident(brand) + "|" + ident(name)


def main():
    if not os.path.exists(SRC):
        sys.exit("no such workbook: " + SRC)

    existing = []
    if os.path.exists(OUT):
        try:
            existing = json.load(io.open(OUT, encoding="utf8"))
        except Exception as e:
            sys.exit("data/raw/xlsx_models.json is unreadable (%s). Move it aside to reimport from scratch." % e)

    store = {}          # key -> record, in insertion order
    for rec in existing:
        store[key_of(rec.get("brand"), rec.get("name"))] = dict(rec)
    started_with = len(store)

    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

    report = {
        "source": os.path.basename(SRC),
        "sheetsInWorkbook": len(wb.sheetnames),
        "sheetsProcessed": 0,
        "recordsBefore": started_with,
        "rowsRead": 0,
        "created": 0,
        "updated": 0,
        "unchanged": 0,
        "skippedNoTitle": 0,
        "duplicateRowsInWorkbook": 0,
        "emptyCellsIgnored": 0,
        "brands": [],
        "unmappedHeaders": [],
        "sheetsWithMissingColumns": [],
    }

    seen_this_run = set()

    for ws in wb.worksheets:
        brand = str(ws.title).strip()
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        cols = map_headers(header)

        if REQUIRED not in cols:
            report["sheetsWithMissingColumns"].append(
                {"sheet": brand, "missing": REQUIRED, "header": [str(h) for h in (header or ())]})
            continue

        for idx, cell in enumerate(header or ()):
            key = loose(cell)
            if key and not any(cols.get(f) == idx for f in cols):
                label = text(cell)
                if label and label not in report["unmappedHeaders"]:
                    report["unmappedHeaders"].append(label)

        report["sheetsProcessed"] += 1
        stats = {"brand": brand, "rows": 0, "created": 0, "updated": 0,
                 "unchanged": 0, "skipped": 0, "duplicates": 0}

        for row in rows:
            if row is None or not any(c is not None and str(c).strip() != "" for c in row):
                continue
            report["rowsRead"] += 1
            stats["rows"] += 1

            name = text(row[cols["name"]]) if cols["name"] < len(row) else None
            if not name:
                report["skippedNoTitle"] += 1
                stats["skipped"] += 1
                continue

            k = key_of(brand, name)
            if k in seen_this_run:
                report["duplicateRowsInWorkbook"] += 1
                stats["duplicates"] += 1
                # still merged below: a later row may carry a value the first lacked
            seen_this_run.add(k)

            incoming = {}
            for field, ci in cols.items():
                raw = row[ci] if ci < len(row) else None
                val = cast(field, raw)
                if val is None:
                    report["emptyCellsIgnored"] += 1
                    continue
                incoming[field] = val

            prior = store.get(k)
            if prior is None:
                rec = {"brand": brand, "sheet": brand}
                rec.update(incoming)
                rec.setdefault("name", name)
                store[k] = rec
                report["created"] += 1
                stats["created"] += 1
                continue

            # Merge. Only real values land; an absent cell leaves what is stored.
            changed = False
            for field, val in incoming.items():
                if prior.get(field) != val:
                    prior[field] = val
                    changed = True
            # The title keeps its stored spelling unless this row differs, in
            # which case the workbook is the source of truth for it.
            if prior.get("name") != name:
                prior["name"] = name
                changed = True
            prior["brand"] = brand
            prior["sheet"] = brand

            if changed:
                report["updated"] += 1
                stats["updated"] += 1
            else:
                report["unchanged"] += 1
                stats["unchanged"] += 1

        report["brands"].append(stats)

    records = list(store.values())
    report["recordsAfter"] = len(records)
    report["duplicatesPrevented"] = report["unchanged"] + report["updated"] + report["duplicateRowsInWorkbook"]

    # Coverage, so a field that is empty everywhere is visible rather than assumed.
    coverage = {}
    for f in COLUMNS:
        coverage[f] = sum(1 for r in records if r.get(f) is not None)
    report["fieldCoverage"] = coverage

    if not DRY:
        os.makedirs(os.path.dirname(OUT), exist_ok=True)
        json.dump(records, io.open(OUT, "w", encoding="utf8"), ensure_ascii=False)
        json.dump(report, io.open(REPORT, "w", encoding="utf8"), ensure_ascii=False, indent=2)

    # ------------------------------------------------------------- summary
    w = sys.stdout.write
    w("\n  Mobile Parts Finder - Excel import%s\n" % ("  (DRY RUN - nothing written)" if DRY else ""))
    w("  " + "-" * 62 + "\n")
    w("  source                 %s\n" % report["source"])
    w("  sheets in workbook     %d\n" % report["sheetsInWorkbook"])
    w("  sheets processed       %d\n" % report["sheetsProcessed"])
    w("  rows read              %d\n" % report["rowsRead"])
    w("  " + "-" * 62 + "\n")
    w("  records before         %d\n" % report["recordsBefore"])
    w("  created                %d\n" % report["created"])
    w("  updated                %d\n" % report["updated"])
    w("  unchanged              %d\n" % report["unchanged"])
    w("  records after          %d\n" % report["recordsAfter"])
    w("  " + "-" * 62 + "\n")
    w("  skipped (no title)     %d\n" % report["skippedNoTitle"])
    w("  duplicate rows in file %d\n" % report["duplicateRowsInWorkbook"])
    w("  empty cells ignored    %d   (left absent, never zero-filled)\n" % report["emptyCellsIgnored"])
    if report["unmappedHeaders"]:
        w("  columns not imported   %s\n" % ", ".join(report["unmappedHeaders"]))
    if report["sheetsWithMissingColumns"]:
        w("  SHEETS SKIPPED         %s\n" %
          ", ".join(s["sheet"] for s in report["sheetsWithMissingColumns"]))

    w("\n  field coverage (of %d records)\n" % report["recordsAfter"])
    for f in COLUMNS:
        n = coverage[f]
        pct = (100.0 * n / report["recordsAfter"]) if report["recordsAfter"] else 0
        w("    %-9s %6d  %5.1f%%%s\n" % (f, n, pct, "" if n else "   <- empty in this workbook"))

    w("\n  per brand\n")
    for s in report["brands"]:
        w("    %-10s rows %5d   new %5d   updated %5d   same %5d%s\n" %
          (s["brand"], s["rows"], s["created"], s["updated"], s["unchanged"],
           ("   skipped %d" % s["skipped"]) if s["skipped"] else ""))

    if not DRY:
        w("\n  wrote %s\n" % os.path.relpath(OUT, ROOT))
        w("  report %s\n" % os.path.relpath(REPORT, ROOT))
        w("\n  next:  node scripts/build-dataset.js && node scripts/build-runtime-bundle.js\n\n")
    else:
        w("\n  dry run - nothing written\n\n")


if __name__ == "__main__":
    main()
